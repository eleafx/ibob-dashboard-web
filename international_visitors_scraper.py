"""
PartnerNet HKTB — International Visitor Arrivals Scraper
=========================================================
Logs into PartnerNet, calls the internal JSON API to fetch monthly visitor
arrivals by market, and writes the results back into the "Copied from website"
sheet in Macro Update_IBOB Master.xlsx.

Usage
-----
    python international_visitors_scraper.py [--year 2026] [--excel path/to/Macro.xlsx]

Requirements
------------
    pip install playwright openpyxl
    playwright install chromium

The script uses Playwright (headless Chromium) so that:
  1. It can perform a proper SSO login (cookie / SAML based).
  2. It can call the credentialled API via the browser's authenticated session.

All 21 markets are scraped including Middle East (API value verified to match Excel).
"""

import argparse
import asyncio
import json
import re
import datetime
from pathlib import Path
import os

# Load .env file if present (so users don't need to export env vars manually)
_env_file = Path(__file__).parent / ".env"
if _env_file.exists():
    for _line in _env_file.read_text().splitlines():
        _line = _line.strip()
        if _line and not _line.startswith("#") and "=" in _line:
            _k, _v = _line.split("=", 1)
            os.environ.setdefault(_k.strip(), _v.strip())

# ---------------------------------------------------------------------------
# Market mapping: PartnerNet API key  →  Excel column header
# ---------------------------------------------------------------------------
MARKET_MAP = {
    "Australia":           "Australia",
    "Canada":              "Canada",
    "France":              "France",
    "Germany":             "Germany",
    "India":               "India",
    "Indonesia":           "Indonesia",
    "Japan":               "Japan",
    "Macau SAR / Not Identified": "Macau SAR",
    "Mainland China":      "Mainland",
    "Malaysia":            "Malaysia",
    "Netherlands":         "Netherlands",
    "Philippines":         "Philippines",
    "Russia":              "Russia",
    "Singapore":           "Singapore",
    "South Korea":         "South Korea",
    "Taiwan":              "Taiwan",
    "Thailand":            "Thailand",
    "United Kingdom":      "United Kingdom",
    "USA":                 "USA",
    "Vietnam":             "Vietnam",
    "Middle East":         "Middle East",   # ✅ API value verified to match Excel exactly
}

# Column order in the Excel "Copied from website" sheet (matches row 1 headers)
EXCEL_COLUMNS = [
    "Date",
    "Australia", "Canada", "France", "Germany", "India", "Indonesia",
    "Japan", "Macau SAR", "Mainland", "Malaysia", "Netherlands",
    "Philippines", "Russia", "Singapore", "South Korea", "Taiwan",
    "Thailand", "United Kingdom", "USA", "Vietnam", "Middle East",
]

# Month name → number
MONTH_NAMES = {
    "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
    "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12,
}


# ---------------------------------------------------------------------------
# Scraper
# ---------------------------------------------------------------------------

async def fetch_visitor_data(username: str, password: str, target_year: int):
    """
    Log into PartnerNet and fetch visitor arrival data for `target_year`.
    Returns dict: {month_int: {market_excel_col: int_value, ...}, ...}
    """
    from playwright.async_api import async_playwright

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context()
        page = await context.new_page()

        # ----------------------------------------------------------------
        # Step 1: Navigate to the tourism performance page (triggers login)
        # ----------------------------------------------------------------
        print("[1/4] Navigating to PartnerNet...")
        await page.goto(
            "https://partnernet.hktb.com/en/research_statistics/tourism_performance/index.html",
            wait_until="networkidle",
            timeout=30_000,
        )

        # ----------------------------------------------------------------
        # Step 2: Login if redirect to SSO
        # ----------------------------------------------------------------
        if "hktbsso" in page.url or "login" in page.url.lower():
            print("[2/4] Logging in via SSO...")
            await page.fill('input[type="text"]', username)
            await page.fill('input[type="password"]', password)
            await page.click('button[type="submit"]')
            await page.wait_for_url(
                "**/partnernet.hktb.com/**",
                timeout=20_000,
            )
            print("      ✓ Logged in")
        else:
            print("[2/4] Already logged in (session active)")

        # ----------------------------------------------------------------
        # Step 3: Call the date-range discovery API
        # ----------------------------------------------------------------
        print("[3/4] Fetching available data range...")
        date_range_resp = await page.evaluate("""
            async () => {
                const r = await fetch('/visitorStat/GetVisitorStatDate?type=data', {cache: 'no-cache'});
                return r.json();
            }
        """)
        print(f"      Data available: "
              f"{date_range_resp['fromYear']}-{date_range_resp['fromMonth']:02d} "
              f"→ {date_range_resp['toYear']}-{date_range_resp['toMonth']:02d}")

        # ----------------------------------------------------------------
        # Step 4: Fetch monthly data for the target year
        # ----------------------------------------------------------------
        to_year  = date_range_resp["toYear"]
        to_month = date_range_resp["toMonth"]

        # Clamp: only fetch months that are available
        if target_year > to_year:
            raise ValueError(
                f"Requested year {target_year} but latest available is {to_year}-{to_month:02d}"
            )

        end_month = to_month if target_year == to_year else 12

        print(f"[4/4] Downloading {target_year} Jan–{end_month:02d} data...")
        data_resp = await page.evaluate(f"""
            async () => {{
                const r = await fetch(
                    '/visitorStat/GetVisitorStat?type=data&includeQuarter=false' +
                    '&fromyear={target_year}&frommonth=1' +
                    '&toyear={target_year}&tomonth={end_month}',
                    {{cache: 'no-cache'}}
                );
                return r.json();
            }}
        """)

        if not data_resp.get("status"):
            raise RuntimeError("API returned status=false. Check login or date range.")

        await browser.close()

        # ----------------------------------------------------------------
        # Parse: extract "COR - Arrivals - Total" for each market per month
        # ----------------------------------------------------------------
        year_data = data_resp["data"].get(str(target_year), {})
        result = {}  # {month_int: {excel_col: int_value}}

        for month_str, markets in year_data.items():
            try:
                month_int = int(month_str)
            except ValueError:
                continue  # skip "Q1", "FY" etc.

            row = {}
            for api_name, excel_col in MARKET_MAP.items():
                market_data = markets.get(api_name, {})
                raw = market_data.get("COR - Arrivals - Total", "")
                if raw:
                    # Remove commas and parse
                    try:
                        row[excel_col] = int(str(raw).replace(",", "").strip())
                    except ValueError:
                        row[excel_col] = None
                else:
                    row[excel_col] = None

            result[month_int] = row
            print(f"      Month {month_int:02d}: "
                  f"{sum(1 for v in row.values() if v is not None)}/{len(MARKET_MAP)} markets OK")

        return result, end_month


# ---------------------------------------------------------------------------
# Excel writer
# ---------------------------------------------------------------------------

def update_excel(excel_path: str, year: int, scraped: dict, end_month: int):
    """
    Write scraped data into the "Copied from website" sheet.
    Row layout: Row 1 = headers, Row 2 = Jan, Row 3 = Feb, ...
    Middle East column is preserved as-is (manual entry).
    """
    import openpyxl
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(excel_path)
    if "Copied from website" not in wb.sheetnames:
        raise KeyError("Sheet 'Copied from website' not found in workbook")

    ws = wb["Copied from website"]

    # Read header row (row 1) to build col_index map
    header_row = [ws.cell(1, c).value for c in range(1, ws.max_column + 2)]
    col_idx = {}  # excel_col_name → 1-based column index
    for i, h in enumerate(header_row, start=1):
        if h:
            col_idx[str(h).strip()] = i

    print(f"\nWriting to Excel: {excel_path}")
    print(f"Columns found: {list(col_idx.keys())}")

    days_in_month = {
        1:31, 2:28, 3:31, 4:30, 5:31, 6:30,
        7:31, 8:31, 9:30, 10:31, 11:30, 12:31,
    }
    # Leap year check
    if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0):
        days_in_month[2] = 29

    for month_int in range(1, end_month + 1):
        excel_row = month_int + 1  # Row 2 = Jan, Row 3 = Feb ...

        month_data = scraped.get(month_int, {})
        if not month_data:
            print(f"  Skipping month {month_int:02d} — no data")
            continue

        # Write Date cell (last day of month as convention)
        date_col = col_idx.get("Date")
        if date_col:
            last_day = days_in_month[month_int]
            ws.cell(excel_row, date_col).value = datetime.datetime(year, month_int, last_day)

        # Write market values
        updates = 0
        for excel_col, value in month_data.items():
            ci = col_idx.get(excel_col)
            if ci and value is not None:
                ws.cell(excel_row, ci).value = value
                updates += 1

        print(f"  Month {month_int:02d}: wrote {updates} market values to row {excel_row}")

    wb.save(excel_path)
    print(f"\n✓ Saved: {excel_path}")
    print("⚠️  Remember to manually enter Middle East values in column V!")


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Scrape PartnerNet visitor arrivals → CSV (GitHub Actions) or Excel (local)"
    )
    parser.add_argument(
        "--year", type=int, default=datetime.datetime.now().year,
        help="Target year (default: current year)",
    )
    parser.add_argument(
        "--output-csv", default=None,
        help="Save as CSV to this path, e.g. data/international_visitors.csv (GitHub Actions mode)",
    )
    parser.add_argument(
        "--excel", default=None,
        help="Path to Macro Update_IBOB Master.xlsx (local mode, auto-detected if omitted)",
    )
    parser.add_argument(
        "--username", default=os.environ.get("PARTNERNET_USER", ""),
        help="PartnerNet login email (or set PARTNERNET_USER env var / .env file)",
    )
    parser.add_argument(
        "--password", default=os.environ.get("PARTNERNET_PASS", ""),
        help="PartnerNet password (or set PARTNERNET_PASS env var / .env file)",
    )
    parser.add_argument("--dry-run", action="store_true", help="Print data without writing output")
    args = parser.parse_args()

    # ----------------------------------------------------------------
    # Credential check
    # ----------------------------------------------------------------
    if not args.username or not args.password:
        print("❌  Missing credentials. Options:")
        print("    1. Copy .env.example → .env and fill in PARTNERNET_USER / PARTNERNET_PASS")
        print("    2. Set env vars: export PARTNERNET_USER=... PARTNERNET_PASS=...")
        print("    3. Pass flags: --username ... --password ...")
        return

    # ----------------------------------------------------------------
    # Auto-detect Excel (only in local mode when --output-csv not given)
    # ----------------------------------------------------------------
    if args.output_csv is None and not args.dry_run and args.excel is None:
        candidates = list(Path.home().glob("Downloads/Macro Update_IBOB Master.xlsx"))
        if not candidates:
            candidates = list(Path.home().glob("**/Macro Update_IBOB Master.xlsx"))
        if candidates:
            args.excel = str(candidates[0])
            print(f"Auto-detected Excel: {args.excel}")
        else:
            print("⚠️  Could not auto-detect Excel. Use --excel or --output-csv to specify output.")
            return

    print(f"Target year : {args.year}")
    print(f"Output CSV  : {args.output_csv or '(none)'}")
    print(f"Excel path  : {args.excel or '(none)'}")
    print(f"Dry run     : {args.dry_run}")
    print()

    # Run async scraper
    scraped, end_month = asyncio.run(
        fetch_visitor_data(args.username, args.password, args.year)
    )

    if args.dry_run:
        print("\n--- DRY RUN: Scraped data (month → market → value) ---")
        for m, row in sorted(scraped.items()):
            print(f"  {args.year}-{m:02d}: {row}")
        return

    # ----------------------------------------------------------------
    # Output: CSV (GitHub Actions) or Excel (local), or both
    # ----------------------------------------------------------------
    if args.output_csv:
        import csv
        csv_path = Path(args.output_csv)
        csv_path.parent.mkdir(parents=True, exist_ok=True)
        markets = list(MARKET_MAP.values())
        fieldnames = ["year", "month"] + markets

        # Merge with existing CSV so prior years remain available for YoY comparisons
        existing_rows = []
        if csv_path.exists():
            with open(csv_path, newline="", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                existing_rows = [
                    row for row in reader
                    if str(row.get("year")) != str(args.year)
                ]

        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(existing_rows)
            for month in sorted(scraped):
                row_data = {"year": args.year, "month": month}
                row_data.update({m: scraped[month].get(m, "") for m in markets})
                writer.writerow(row_data)
        print(f"✓ CSV saved: {args.output_csv}")

    if args.excel:
        update_excel(args.excel, args.year, scraped, end_month)


if __name__ == "__main__":
    main()
